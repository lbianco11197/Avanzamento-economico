# app.py — Avanzamento: GitHub + filtro mese globale + filtro tecnico + invio mese scelto
import io
import os
import time
import base64
import requests
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
import smtplib, re
from email.mime.text import MIMEText
from pandas import ExcelWriter

# =========================
# CONFIG (coerente con il tuo .py precedente)
# =========================
REPO_OWNER = "lbianco11197"
REPO_NAME  = "Avanzamento-economico"
BRANCH     = "main"
XLSX_PATH  = "Avanzamento.xlsx"      # percorso nel repo
SHEET_NAME = ""                      # "" => primo foglio
PAGE_TITLE = "Avanzamento mensile €/h per Tecnico - Euroirte s.r.l."

# GitHub token (opzionale, per rate-limit migliori)
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", os.getenv("GITHUB_TOKEN"))

API_URL     = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{XLSX_PATH}"
COMMITS_API = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/commits"

# SMTP (come nel vecchio script)
SMTP_HOST    = "mail.euroirte.it"
MAIL_SUBJECT = "EUROIRTE - Avanzamento Economico"
SMTP_USER    = str(st.secrets["SMTP_USER"]).strip()
SMTP_PASS    = str(st.secrets["SMTP_PASS"]).strip()
SMTP_FROM    = str(st.secrets.get("SMTP_FROM", SMTP_USER)).strip()
EMAIL_RE     = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Mesi in italiano
MESI_IT = ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
           "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

# =========================
# PAGE
# =========================
st.set_page_config(layout="wide", page_title=PAGE_TITLE, page_icon=":bar_chart:")
st.title(f"📊 {PAGE_TITLE}")

# --- STILE: bordi grigio chiaro per select/menu/input ---
st.markdown("""
<style>
  /* Contenitori input/table */
  .stTextInput, .stNumberInput, .stDateInput, .stMultiSelect, .stRadio,
  .stCheckbox, .stSlider, .stFileUploader, .stTextArea, .stSelectbox {
    background-color: rgba(255,255,255,0.88) !important;
    border-radius: 10px !important;
    border: 1px solid #ddd !important;
  }
  /* Selectbox (menu a tendina) con bordo #ddd */
  .stSelectbox div[data-baseweb="select"] {
    background-color: rgba(255,255,255,0.88) !important;
    border-radius: 10px !important;
    border: 1px solid #ddd !important;
  }
  /* Multiselect coerente */
  .stMultiSelect div[data-baseweb="select"] {
    background-color: rgba(255,255,255,0.88) !important;
    border-radius: 10px !important;
    border: 1px solid #ddd !important;
  }
</style>
""", unsafe_allow_html=True)

# =========================
# HELPERS: GitHub fetch
# =========================
def _headers():
    h = {"Accept": "application/vnd.github+json"}
    if GITHUB_TOKEN:
        h["Authorization"] = f"token {GITHUB_TOKEN}"
    return h

@st.cache_data(show_spinner=True, ttl=60)
def fetch_excel_bytes_via_api():
    """
    Legge Avanzamento.xlsx dall'API GitHub.
    Ritorna (sha, bytes, last_modified_human).
    """
    r = requests.get(API_URL, headers=_headers(), params={"ref": BRANCH}, timeout=30)
    r.raise_for_status()
    j = r.json()

    # data ultimo commit sul file (umano)
    r2 = requests.get(COMMITS_API, headers=_headers(),
                      params={"path": XLSX_PATH, "per_page": 1, "sha": BRANCH}, timeout=30)
    last_human = None
    if r2.ok:
        lst = r2.json()
        if isinstance(lst, list) and lst:
            iso = lst[0]["commit"]["committer"]["date"]
            try:
                dt = datetime.fromisoformat(iso.replace("Z", "+02:00"))
                last_human = dt.strftime("%d/%m/%Y")
            except Exception:
                last_human = iso

    # contenuto inline (base64) oppure download_url
    content = j.get("content"); encoding = j.get("encoding")
    if content and encoding == "base64":
        data = base64.b64decode(content)
        return j.get("sha") or str(int(time.time())), data, last_human

    download_url = j.get("download_url")
    if download_url:
        r3 = requests.get(download_url, timeout=30, headers={"Cache-Control": "no-cache"})
        r3.raise_for_status()
        return j.get("sha") or str(int(time.time())), r3.content, last_human

    raise RuntimeError("Impossibile ottenere il contenuto del file da GitHub.")

def load_avanzamento_df_from_bytes(xls_bytes: bytes) -> pd.DataFrame:
    """
    Parsa l'Excel (valori calcolati) e mantiene:
    - 'Tecnico' (str ripulito)
    - 'Data aggiornamento' (datetime per calcoli)
    - 'Data Aggiornamento' (stringa gg/mm/aaaa per tabella/email)
    - 'Ore lavorate' (float)
    - 'Avanzamento €/h' (float, 2 decimali in visualizzazione)
    - 'Mail' (str)
    """
    bio = io.BytesIO(xls_bytes)
    wb = load_workbook(bio, data_only=True, read_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME and SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(columns=[
            "Tecnico","Data aggiornamento","Data Aggiornamento","Ore lavorate","Avanzamento €/h","Mail"
        ])

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header)

    # normalizza nomi più comuni
    rename_map = {}
    for c in df.columns:
        key = str(c).strip().lower().replace(" ", "")
        if key.startswith("tecnico"):
            rename_map[c] = "Tecnico"
        elif key.startswith("data"):
            rename_map[c] = "Data aggiornamento"
        elif key in ("orelavorate","ore","orelavorate(h)"):
            rename_map[c] = "Ore lavorate"
        elif key in ("avanzamento€/h","avanzamento","avanzamentoeuro/ora","avanzamentoeuroh"):
            rename_map[c] = "Avanzamento €/h"
        elif key in ("mail","email","e-mail"):
            rename_map[c] = "Mail"
    df = df.rename(columns=rename_map)

    wanted = ["Tecnico", "Data aggiornamento", "Ore lavorate", "Avanzamento €/h", "Mail"]
    for w in wanted:
        if w not in df.columns:
            df[w] = None
    df = df[wanted].copy()

    # tipi
    df["Ore lavorate"] = pd.to_numeric(df["Ore lavorate"], errors="coerce")
    df["Avanzamento €/h"] = pd.to_numeric(df["Avanzamento €/h"], errors="coerce")

    # datetime per calcoli + stringa per mostrata
    df["Data aggiornamento"] = pd.to_datetime(df["Data aggiornamento"], errors="coerce", dayfirst=True)
    df["Data Aggiornamento"] = df["Data aggiornamento"].dt.strftime("%d/%m/%Y")

    # pulizia tecnici
    df = df.dropna(how="all")
    if "Tecnico" in df.columns:
        df["Tecnico"] = (
            df["Tecnico"].astype(str).fillna("")
            .str.strip().str.replace(r"\s+", " ", regex=True)
        )
        df = df[df["Tecnico"] != ""]
    return df.reset_index(drop=True)

# =========================
# CARICAMENTO da GitHub
# =========================
cols = st.columns([0.2, 0.8])
with cols[0]:
    if st.button("🔄 Aggiorna dati"):
        st.cache_data.clear()

try:
    version_sha, xls_bytes, last_update_date = fetch_excel_bytes_via_api()
    df = load_avanzamento_df_from_bytes(xls_bytes)
except Exception as e:
    st.error(f"Errore nel caricamento da GitHub: {e}")
    st.stop()

if last_update_date:
    st.caption(f"📅 Dati aggiornati al {last_update_date}")

# =========================
# MESE GLOBALE (mostra nomi mese)
# =========================
df["Mese"] = df["Data aggiornamento"].dt.to_period("M").dt.to_timestamp()

def nome_mese_it(ts: pd.Timestamp) -> str:
    # es. Timestamp('2025-08-01') -> "Agosto"
    if pd.isna(ts):
        return ""
    return MESI_IT[int(ts.month)]

mesi_disponibili = sorted(df["Mese"].dropna().unique())
if not mesi_disponibili:
    st.warning("Nessuna data valida trovata.")
    st.stop()

col_mese, col_toggle = st.columns([3, 2])
with col_mese:
    mese_scelto = st.selectbox(
        "📅 Scegli il mese (filtro globale)",
        mesi_disponibili,
        index=len(mesi_disponibili)-1,
        format_func=nome_mese_it,   # <<< NOME MESE
        key="mese_globale"
    )
with col_toggle:
    mostra_tutti = st.toggle("Mostra tutti i mesi", value=False)

if not mostra_tutti:
    df = df[df["Mese"] == mese_scelto].copy()

# =========================
# FILTRO SINGOLO TECNICO
# =========================
st.divider()
st.subheader("🔎 Filtri")
tecnici = sorted(df["Tecnico"].dropna().unique().tolist())
opzioni_tec = ["— Tutti i tecnici —"] + tecnici
tec_scelto = st.selectbox("Tecnico", opzioni_tec, index=0)

if tec_scelto != "— Tutti i tecnici —":
    df = df[df["Tecnico"] == tec_scelto].copy()

# =========================
# TABELLINA CON LOGICA SEMAFORICA (€/h 2 decimali)
# =========================
st.subheader("📋 Dati correnti")
def style_semaforo(val):
    """Rosso <30, Giallo 30–35, Verde >35."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return ""
    if v < 30:
        return "background-color: #ff4d4d;"   # rosso
    elif 30 <= v <= 35:
        return "background-color: #ffff99;"   # giallo
    else:
        return "background-color: #b3ffb3;"   # verde

preview_cols = ["Tecnico", "Data Aggiornamento", "Ore lavorate", "Avanzamento €/h", "Mail"]
df_preview = df[preview_cols].copy()

styler = (
    df_preview.style
    .format({
        "Ore lavorate": "{:.0f}",
        "Avanzamento €/h": "€{:.2f}/h",   # <<< due decimali
    })
    .applymap(style_semaforo, subset=["Avanzamento €/h"])  # <<< semaforo su €/h
)

st.table(styler)  # st.table preserva lo Styler (st.dataframe lo ignora)

# =========================
# EXPORT vista corrente (facoltativo)
# =========================
def _to_excel_bytes(frame: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with ExcelWriter(buf, engine="openpyxl") as w:
        frame.to_excel(w, index=False, sheet_name="Avanzamento")
    buf.seek(0)
    return buf.getvalue()

st.download_button(
    "⬇️ Esporta vista corrente (xlsx)",
    data=_to_excel_bytes(df_preview),
    file_name="avanzamento_corrente.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

# =========================
# SMTP: TEST CONNESSIONE (come prima)
# =========================
st.divider()
st.subheader("🔍 Test connessione SMTP")
if SMTP_FROM.lower() != SMTP_USER.lower():
    st.warning(f"Mittente ({SMTP_FROM}) diverso dall'utente SMTP ({SMTP_USER}). Alcuni server lo vietano.")

def _smtp_login_test():
    try:
        with smtplib.SMTP_SSL(SMTP_HOST, 465, timeout=20) as s:
            s.noop()
            s.login(SMTP_USER, SMTP_PASS)
        return True, "Login OK via SSL (465)"
    except Exception as e_ssl:
        try:
            with smtplib.SMTP(SMTP_HOST, 587, timeout=20) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(SMTP_USER, SMTP_PASS)
            return True, "Login OK via STARTTLS (587)"
        except Exception as e_tls:
            return False, f"Errore login.\nSSL465: {e_ssl}\nSTARTTLS587: {e_tls}"

if st.button("Esegui test SMTP"):
    ok, msg = _smtp_login_test()
    (st.success if ok else st.error)(msg)

# =========================
# INVIO EMAIL — SCELTA MESE (solo dati di quel mese, ignorando filtri attuali)
# =========================
st.divider()
st.subheader("📧 Invia email")

# per l'invio usiamo l'intero dataset (no filtri), ma mostriamo i mesi con nome
# ricarichiamo dalla stessa sorgente per essere certi di non avere filtri pendenti
_, xls_bytes_for_email, _ = fetch_excel_bytes_via_api()
df_all = load_avanzamento_df_from_bytes(xls_bytes_for_email)
df_all["Mese"] = df_all["Data aggiornamento"].dt.to_period("M").dt.to_timestamp()
mesi_email = sorted(df_all["Mese"].dropna().unique(), reverse=True)

# default = mese globale
default_idx = mesi_email.index(mese_scelto) if mese_scelto in mesi_email else 0
mese_email = st.selectbox(
    "Mese da inviare",
    mesi_email,
    index=default_idx,
    format_func=lambda x: nome_mese_it(x),   # nome mese
    key="mese_da_inviare"
)

if st.button("✉️ Invia email per il mese selezionato"):
    df_email = df_all[df_all["Mese"] == mese_email].copy()
    if df_email.empty:
        st.warning("Nessun dato per il mese selezionato.")
        st.stop()

    risultati, invalidi = [], []
    try:
        # tenta SSL465, poi fallback 587
        try:
            server = smtplib.SMTP_SSL(SMTP_HOST, 465, timeout=30)
            mode = "SSL465"
        except Exception:
            server = smtplib.SMTP(SMTP_HOST, 587, timeout=30)
            server.ehlo(); server.starttls(); server.ehlo()
            mode = "STARTTLS587"

        with server:
            server.login(SMTP_USER, SMTP_PASS)
            for _, r in df_email.iterrows():
                nome = str(r.get("Tecnico", "")).strip()
                to = ("" if pd.isna(r.get("Mail")) else str(r.get("Mail"))).strip()

                if not to or not EMAIL_RE.match(to):
                    invalidi.append((nome, to))
                    continue

                data  = r.get("Data Aggiornamento", "")
                avanz = float(pd.to_numeric(r.get("Avanzamento €/h"), errors="coerce") or 0)
                ore   = float(pd.to_numeric(r.get("Ore lavorate"), errors="coerce") or 0)

                # corpo con €/h a 2 decimali
                corpo = (
                    f"Ciao {nome},\n\n"
                    f"il tuo avanzamento economico aggiornato al {data} è di {avanz:.2f} €/h "
                    f"e il totale delle ore lavorate è {ore:.0f}.\n"
                )

                msg = MIMEText(corpo, "plain", "utf-8")
                msg["Subject"] = MAIL_SUBJECT
                msg["From"] = SMTP_FROM
                msg["To"] = to

                refused = server.send_message(msg)
                if refused:
                    risultati.append(("❌", nome, to))
                else:
                    risultati.append(("✅", nome, to))

        st.success(f"Invio completato ({mode})")
        st.dataframe(pd.DataFrame(risultati, columns=["Stato", "Tecnico", "Email"]), use_container_width=True)
        if invalidi:
            st.warning("Email non valide/assenti")
            st.dataframe(pd.DataFrame(invalidi, columns=["Tecnico", "Email"]), use_container_width=True)

    except smtplib.SMTPAuthenticationError as e:
        st.error(f"Autenticazione fallita: {e}")
    except Exception as e:
        st.error(f"Errore durante l'invio: {e}")
